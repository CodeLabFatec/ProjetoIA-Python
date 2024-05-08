import io
import time
from datetime import *
from flask import send_file
from openpyxl import Workbook
from openpyxl import load_workbook
import os

from repository.dados_ia import informacoesRepository

class ExportarRelatorio:


    def export_relatorio_7_dias(template_filename):

        current_dir = os.path.dirname(os.path.abspath(__file__))
        template_path = os.path.abspath(os.path.join(current_dir, '..', '..', '..', 'assets', template_filename))

        wb = load_workbook(template_path)
        sheet = wb.active

        data = informacoesRepository.calcula_entrada_pessoas()
        data_inicial = datetime.now().date()
        while data_inicial.strftime("%A") != "Monday":
            data_inicial -= timedelta(days=1)

        for i, entry in enumerate(data[:7], start=3):
            data_entrada_formatada = datetime.strptime(entry["data_entrada"], "%Y-%m-%d").strftime("%d/%m/%Y") if entry["data_entrada"] else ""
            sheet[f"C{i}"] = data_entrada_formatada
            sheet[f"D{i}"] = entry["numero_pessoas"]

        if len(data) < 7:
            for i in range(len(data) + 3, 10):
                sheet[f"C{i}"] = ""
                sheet[f"D{i}"] = ""
                sheet[f"B{i}"] = ""

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        file_name = f"Relatorio_7_dias.xlsx"
        response = send_file(
            output,
            as_attachment=True,
            download_name=file_name,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response.headers['Content-Disposition'] = f'attachment; filename={file_name}'
        return response

    def export_relatorio_14_dias(template_filename):

        current_dir = os.path.dirname(os.path.abspath(__file__))
        template_path = os.path.abspath(os.path.join(current_dir, '..', '..', '..', 'assets', template_filename))

        wb = load_workbook(template_path)
        sheet = wb.active

        data = informacoesRepository.calcula_entrada_pessoas()
        data_inicial = datetime.now().date()
        while data_inicial.strftime("%A") != "Monday":
            data_inicial -= timedelta(days=1)

        for i, entry in enumerate(data[:14], start=3):
            data_entrada_formatada = datetime.strptime(entry["data_entrada"], "%Y-%m-%d").strftime("%d/%m/%Y") if entry["data_entrada"] else ""
            sheet[f"C{i}"] = data_entrada_formatada
            sheet[f"D{i}"] = entry["numero_pessoas"]

        if len(data) < 14:
            for i in range(len(data) + 3, 17):
                sheet[f"C{i}"] = ""
                sheet[f"D{i}"] = ""
                sheet[f"B{i}"] = ""

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        file_name = f"Relatorio_14_dias.xlsx"
        response = send_file(
            output,
            as_attachment=True,
            download_name=file_name,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response.headers['Content-Disposition'] = f'attachment; filename={file_name}'
        return response
    
    def generate_redzone_log_excel(id):

        wb = Workbook()
        sheet = wb.active

        sheet.append(["Data", "Hora", "Tipo", "Redzone"])

        redzone_entries = informacoesRepository.get_redzone_entries(id)

        for entry in redzone_entries:
            data_datetime = entry["data"]
            tipo = entry["tipo"]
            nome_redzone = entry["nome"]
            data = ""
            hora = ""

            if data_datetime:
                data = data_datetime.strftime("%d/%m/%Y")
                hora = data_datetime.strftime("%H:%M")

            sheet.append([data, hora, tipo, nome_redzone])

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        file_name = f"Redzone_Log_{id}.xlsx"
        response = send_file(
            output,
            as_attachment=True,
            download_name=file_name,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response.headers['Content-Disposition'] = f'attachment; filename={file_name}'
        
        return response

    def generate_redzone_log_excel_all(): 

        wb = Workbook()
        sheet = wb.active

        sheet.append(["Data", "Hora", "Tipo", "Redzone"])

        redzone_entries = informacoesRepository.get_all_redzone_entries()

        for entry in redzone_entries:
            data_datetime = entry["data"]
            tipo = entry["tipo"]
            nome_redzone = entry["nome"]
            data = ""
            hora = ""

            if data_datetime:
                data = data_datetime.strftime("%d/%m/%Y")
                hora = data_datetime.strftime("%H:%M")

            sheet.append([data, hora, tipo, nome_redzone])

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        file_name = f"Redzone_Log.xlsx"
        response = send_file(
            output,
            as_attachment=True,
            download_name=file_name,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response.headers['Content-Disposition'] = f'attachment; filename={file_name}'
        
        return response
    
    def export_relatorio_por_id(template_filename, id):
        current_dir = os.path.dirname(os.path.abspath(__file__))
        template_path = os.path.abspath(os.path.join(current_dir, '..', '..', '..', 'assets', template_filename))

        wb = load_workbook(template_path)
        sheet = wb.active

        data = informacoesRepository.calcula_entrada_pessoas_por_id(id)
        for i, entry in enumerate(data, start=3):
            data_entrada_formatada = entry["data_entrada"].strftime("%d/%m/%Y") if entry["data_entrada"] else ""
            sheet[f"C{i}"] = data_entrada_formatada
            sheet[f"D{i}"] = entry["numero_pessoas"]

        if len(data) < 7:
            for i in range(len(data) + 3, 10):
                sheet[f"C{i}"] = ""
                sheet[f"D{i}"] = ""
                sheet[f"B{i}"] = ""

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        file_name = f"Relatorio_{id}.xlsx"
        response = send_file(
            output,
            as_attachment=True,
            download_name=file_name,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response.headers['Content-Disposition'] = f'attachment; filename={file_name}'
        return response
